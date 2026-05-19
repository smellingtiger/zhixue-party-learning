"""
删除不在Excel中的课程文件
"""

import openpyxl
import re
from pathlib import Path

TXT_DIR = Path(r"E:\社院课程stt\knowledge_base_txt")
EXCEL_FILE = Path(r"d:\TraeProject\zhixue-party-learning\精英课程资源库课程清单（2015年-2026年1月）_已标注.xlsx")

# 读取Excel中的所有课程编码
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb.active
excel_codes = set()
for row in range(2, ws.max_row + 1):
    code = ws.cell(row=row, column=1).value
    if code:
        excel_codes.add(str(code).strip())

print(f"Excel中共有 {len(excel_codes)} 个课程编码")

# 查找并删除不在Excel中的txt文件
txt_files = list(TXT_DIR.glob("*.txt"))
deleted_count = 0

for txt_file in txt_files:
    try:
        content = txt_file.read_text(encoding="utf-8")
        course_code = ""
        for line in content.split("\n")[:6]:
            m = re.search(r"【课程名称】(.+)", line)
            if m:
                course_code = m.group(1).strip()
                break
        
        if course_code and course_code not in excel_codes:
            txt_file.unlink()
            deleted_count += 1
            if deleted_count <= 10 or deleted_count % 50 == 0:
                print(f"已删除: {txt_file.name}")
    except Exception as e:
        print(f"处理 {txt_file.name} 时出错: {e}")

print(f"\n总共删除了 {deleted_count} 个不在Excel中的课程文件")

# 统计剩余
remaining = list(TXT_DIR.glob("*.txt"))
print(f"剩余课程文件: {len(remaining)} 个")
