"""
从Excel文件中提取课程编码和中文名称的映射
"""

import openpyxl
from pathlib import Path

EXCEL_FILE = Path(r"d:\TraeProject\zhixue-party-learning\精英课程资源库课程清单（2015年-2026年1月）_已标注.xlsx")

def extract_course_mapping():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    course_map = {}
    skipped_rows = 0
    
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        
        if code and name:
            code = str(code).strip()
            name = str(name).strip()
            if code and name:
                course_map[code] = name
        else:
            skipped_rows += 1
    
    print(f"从Excel中提取到 {len(course_map)} 个课程编码-名称映射")
    print(f"跳过 {skipped_rows} 行(无编码或名称)")
    
    return course_map

if __name__ == "__main__":
    mapping = extract_course_mapping()
    
    # 打印示例
    print("\n前20个课程映射:")
    for i, (code, name) in enumerate(list(mapping.items())[:20]):
        print(f"  {code} -> {name}")
    
    # 保存到JSON
    import json
    output_file = Path(r"d:\TraeProject\zhixue-party-learning\knowledge-server\course_name_mapping.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)
    print(f"\n映射数据已保存到: {output_file}")
