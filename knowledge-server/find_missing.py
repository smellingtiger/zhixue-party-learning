"""
查找两个Excel文件的区别，找出缺失的课程编码
"""

import openpyxl
from pathlib import Path

# 读取已标注的Excel
wb1 = openpyxl.load_workbook(Path(r"d:\TraeProject\zhixue-party-learning\精英课程资源库课程清单（2015年-2026年1月）_已标注.xlsx"))
ws1 = wb1.active
codes1 = set()
for row in range(2, ws1.max_row + 1):
    code = ws1.cell(row=row, column=1).value
    if code:
        codes1.add(str(code).strip())

# 读取原始Excel
wb2 = openpyxl.load_workbook(Path(r"d:\TraeProject\zhixue-party-learning\精英课程资源库课程清单（2015年-2026年1月）.xlsx"))
ws2 = wb2.active
codes2 = set()
name_map2 = {}
for row in range(2, ws2.max_row + 1):
    code = ws2.cell(row=row, column=1).value
    name = ws2.cell(row=row, column=2).value
    if code:
        code = str(code).strip()
        codes2.add(code)
        if name:
            name_map2[code] = str(name).strip()

# 查找 DSPTXYZY20041710
target = "DSPTXYZY20041710"
print(f"查找: {target}")
print(f"在已标注Excel中: {target in codes1}")
print(f"在原始Excel中: {target in codes2}")
if target in name_map2:
    print(f"原始Excel中的名称: {name_map2[target]}")

# 统计差异
print(f"\n已标注Excel课程数: {len(codes1)}")
print(f"原始Excel课程数: {len(codes2)}")
print(f"原始中有但已标注中没有的课程数: {len(codes2 - codes1)}")

# 查找所有DSPTXYZY开头的课程
dspt_in_1 = [c for c in codes1 if c.startswith('DSPTXYZY')]
dspt_in_2 = [c for c in codes2 if c.startswith('DSPTXYZY')]
print(f"\n已标注中DSPTXYZY开头的课程: {len(dspt_in_1)}")
print(f"原始中DSPTXYZY开头的课程: {len(dspt_in_2)}")
print(f"原始中有但已标注中没有的DSPTXYZY课程:")
for c in sorted(codes2 - codes1):
    if c.startswith('DSPTXYZY'):
        print(f"  {c} -> {name_map2.get(c, '无名称')}")
